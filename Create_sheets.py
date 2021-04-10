from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws1 = wb.create_sheet('Sheet1')

ws2 = wb.create_sheet('Sheet2')

ws3 = wb.create_sheet('Sheet3')

wb.save("create_sheet.xlsx")

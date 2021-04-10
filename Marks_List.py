from openpyxl import Workbook

from openpyxl.styles import Font

wb=Workbook()

wb['Sheet'].title='Marks'

sh1=wb.active

sh1['A1'].value="Name of the class"

sh1['A1'].font=Font(name='Cambria', bold=True, size=14)

sh1.merge_cells('A1:K1')

sh1['A2'].value="Roll No"

sh1['A2'].font=Font(name='Cambria', bold=True, size=12)

sh1.merge_cells('A2:A3')

sh1['B2'].value="Name"

sh1['B2'].font=Font(name='Cambria', bold=True, size=12)

sh1.merge_cells('B2:B3')

sh1['C2'].value="Subjects"

sh1['C2'].font=Font(name='Cambria', bold=True, size=12)

sh1.merge_cells('C2:H2')

sh1['C3'].value="Telugu"

sh1['C3'].font=Font(name='Cambria', bold=True, size=12)

sh1['D3'].value="Hindi"

sh1['D3'].font=Font(name='Cambria', bold=True, size=12)

sh1['E3'].value="English"

sh1['E3'].font=Font(name='Cambria', bold=True, size=12)

sh1['F3'].value="Maths"

sh1['F3'].font=Font(name='Cambria', bold=True, size=12)

sh1['G3'].value="Science"

sh1['G3'].font=Font(name='Cambria', bold=True, size=12)

sh1['H3'].value="Social"

sh1['H3'].font=Font(name='Cambria', bold=True, size=12)

sh1['I2'].value="Total"

sh1['I4']= '=SUM(C4:H4)'

sh1['I2'].font=Font(name='Cambria', bold=True, size=12)

sh1.merge_cells('I2:I3')

sh1['J2'].value="Average"

sh1['J4']= '=AVERAGE(H4/6)'

sh1['J2'].font=Font(name='Cambria', bold=True, size=12)

sh1.merge_cells('J2:J3')

sh1['K2'].value="Rank"

sh1['K4']= '=RANK(I4,I$4:I$40)'

sh1['K2'].font=Font(name='Cambria', bold=True, size=12)

sh1.merge_cells('K2:K3')

wb.save("Marks_List.xlsx")

from openpyxl import Workbook

from openpyxl.styles import Font

wb=Workbook()

wb['Sheet'].title='Time Table'

sh1=wb.active

sh1['A1'].value='Name of the Institution'

sh1.merge_cells('A1:G1')

sh1['A1'].font=Font(name='Cambria', bold=True, size=14)

sh1['A2'].value='Table Table for Course and Acadamic Year'

sh1.merge_cells('A2:G2')

sh1['A2'].font=Font(name='Cambria', bold=True, size=12, underline='single')

sh1['A4'].value='Day/Period'

sh1['A4'].font=Font(name='Cambria', bold=True, size=11)

sh1['A5'].value='Monday'

sh1['A5'].font=Font(name='Cambria', bold=True, size=11)

sh1['A6'].value='Tuesday'

sh1['A6'].font=Font(name='Cambria', bold=True, size=11)

sh1['A7'].value='Wenesday'

sh1['A7'].font=Font(name='Cambria', bold=True, size=11)

sh1['A8'].value='Thursday'

sh1['A8'].font=Font(name='Cambria', bold=True, size=11)

sh1['A9'].value='Friday'

sh1['A9'].font=Font(name='Cambria', bold=True, size=11)

sh1['A10'].value='Saterday'

sh1['A10'].font=Font(name='Cambria', bold=True, size=11)

sh1['B4'].value='10:00-11:00'

sh1['B4'].font=Font(name='Cambria', bold=True, size=11)

sh1['C4'].value='11:00-12:00'

sh1['C4'].font=Font(name='Cambria', bold=True, size=11)

sh1['D4'].value='12:00-01:00'

sh1['D4'].font=Font(name='Cambria', bold=True, size=11)

sh1['E4'].value='01:00-01:30'

sh1['E4'].font=Font(name='Cambria', bold=True, size=11)

sh1['F4'].value='01:30-02:30'

sh1['F4'].font=Font(name='Cambria', bold=True, size=11)

sh1['G4'].value='02:30-03:30'

sh1['G4'].font=Font(name='Cambria', bold=True, size=11)

wb.save("Time_Table.xlsx")
